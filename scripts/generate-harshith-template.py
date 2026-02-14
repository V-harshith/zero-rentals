#!/usr/bin/env python3
"""
Generate Harshith-style Excel template for ZeroRentals property import
Simple format: Sheet1 with headers in row 1, data from row 2 onwards
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook with single sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# Define headers matching Harshith.xlsx exactly
headers = [
    "Country", "City", "Area", "Locality", "PG's for", "Property Name",
    "Owner Name", "Owner Contact", "Landmark", "USP", "Facilities",
    "Private Room", "Double Sharing", "Triple Sharing", "Four Sharing",
    "Deposit", "Address", "PSN", "Owner Contact", "Email", "Embassy tech"
]

# Set column widths (matching typical Excel widths)
widths = [10, 12, 15, 12, 10, 30, 15, 14, 25, 15, 40, 12, 14, 14, 12, 10, 35, 8, 14, 25, 15]
for idx, width in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

# Style for headers
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
header_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Write headers in row 1
for idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Add sample data rows (matching Harshith.xlsx examples)
sample_data = [
    ["India", "Bengaluru", "Bellandur", "Bellandur", "Gents", "PG Ashok gents pg",
     "Ashok", "9663634346", "MAin road 8th cross", "None", "3 times food, House Keeping, Wifi unlimited",
     "None", "9000", "7000", "None", "2000", "https://tinyurl.com/53bts5jd", "155", "9663634346", "9663634346@gmail.com", ""],
    ["India", "Bengaluru", "Bellandur", "Bellandur", "Gents", "Swamy narendra pg gents",
     "ramanaiah", "9620003666", "Ivory studio", "None", "weekdays 2times weekend 3times",
     "7000", "7500", "6000", "None", "3000", "https://tinyurl.com/4ptxpmb4", "84", "9620003666", "9620003666@gmail.com", ""],
    ["India", "Bengaluru", "Bellandur", "Bellandur", "Gents", "RR Gents pg for Ladies",
     "Shiva Reddy", "9620373730", "Dr.Puneet Rajkuma Rd", "", "", 
     "", "", "", "", "", "", "75", "", "", ""],
]

# Write sample data
for row_idx, row_data in enumerate(sample_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=False)

# Save
output_path = r'c:\Users\hxrshith-pc\All_projects\zerorentals\zero-rentals\ZeroRentals_Harshith_Import_Template.xlsx'
wb.save(output_path)
print(f"Template created: {output_path}")

